#!/usr/bin/env python3
"""Copy a metadata workbook and rewrite the Masking of one Summary row.

Used by scripts/masking_sweep.sh to build one workbook per masking variant of a
single (lane, group). Everything else in the workbook — the barcode tabs in
particular — is copied through untouched.

The Summary sheet is read elsewhere with `header=2`, i.e. the header is Excel
row 3 and data starts at row 4. Columns are located by header text, not index.

Usage:
    python src/make_masking_variant_xlsx.py \
        --base metadata/<run>.xlsx --out metadata/<run>_R2-39.xlsx \
        --lane 8 --group 1 --masking "R1:28, I1:10, I2:10, R2:39" \
        --lab-id-suffix R2-39
"""

import argparse
import re
import shutil
import sys

import openpyxl

HEADER_ROW = 3  # Excel row holding the Summary column names


def _as_int(value):
    """Summary cells hold Lane/Gr as int or str depending on the row."""
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def find_columns(ws, names):
    """Map header text -> column index for the requested header names."""
    found = {}
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(HEADER_ROW, col).value or "").strip()
        if header in names:
            found[header] = col
    missing = sorted(names - set(found))
    if missing:
        raise SystemExit(
            f"Summary header row {HEADER_ROW} is missing column(s): {', '.join(missing)}"
        )
    return found


def find_row(ws, cols, lane, group):
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if _as_int(ws.cell(row, cols["Lane"]).value) != lane:
            continue
        if _as_int(ws.cell(row, cols["Gr"]).value) != group:
            continue
        return row
    raise SystemExit(f"No Summary row found for lane {lane} group {group}")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--base", required=True, help="source workbook")
    parser.add_argument("--out", required=True, help="variant workbook to write")
    parser.add_argument("--lane", type=int, required=True)
    parser.add_argument("--group", type=int, required=True)
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--masking", help="replacement Masking string, written verbatim")
    group.add_argument(
        "--set-r2",
        type=int,
        metavar="N",
        help="keep the row's existing Masking but rewrite its R2 term to R2:N. "
             "Preferred over --masking: R1/I1/I2 stay exactly as the run specified them",
    )
    parser.add_argument(
        "--lab-id-suffix",
        default="",
        help="appended to the row's Lab ID as _<suffix>, so the delivered folder "
             "name ({Lab ID}_{Order ID}_{library}_L{lane}_G{group}) is unique per variant",
    )
    args = parser.parse_args()

    shutil.copyfile(args.base, args.out)

    wb = openpyxl.load_workbook(args.out)
    if "Summary" not in wb.sheetnames:
        raise SystemExit(f"{args.base} has no Summary sheet (MiSeq workbook?)")
    ws = wb["Summary"]

    cols = find_columns(ws, {"Lane", "Gr", "Masking", "Lab ID"})
    row = find_row(ws, cols, args.lane, args.group)

    old_masking = ws.cell(row, cols["Masking"]).value
    if args.masking:
        new_masking = args.masking
    else:
        new_masking, subs = re.subn(
            r"(R2\s*:\s*)\d+", rf"\g<1>{args.set_r2}", str(old_masking), count=1
        )
        if subs != 1:
            raise SystemExit(
                f"Masking {old_masking!r} for lane {args.lane} group {args.group} has no "
                f"R2:<n> term to rewrite; pass --masking to set it explicitly"
            )
    ws.cell(row, cols["Masking"]).value = new_masking

    old_lab_id = str(ws.cell(row, cols["Lab ID"]).value or "").strip()
    new_lab_id = old_lab_id
    if args.lab_id_suffix:
        new_lab_id = f"{old_lab_id}_{args.lab_id_suffix}"
        ws.cell(row, cols["Lab ID"]).value = new_lab_id

    wb.save(args.out)

    print(f"{args.out}: lane {args.lane} group {args.group} (Summary row {row})")
    print(f"  Masking: {old_masking!r} -> {new_masking!r}")
    if new_lab_id != old_lab_id:
        print(f"  Lab ID : {old_lab_id!r} -> {new_lab_id!r}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
